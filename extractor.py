"""
MP Bhoj Open University - Result PDF -> Excel extractor
==========================================================
Ye module Crystal Reports se bane MP Bhoj result PDFs (jaise
"RESULT SHEET OF B.SC. FIRST YEAR...") ko parse karta hai.

Column layout (x-position based, points):
    S.No      : x0 < 45
    Roll No   : 45  <= x0 < 155
    Name      : 155 <= x0 < 259
    S/D/W Of  : 259 <= x0 < 371
    Result/Total zone : 371 <= x0 < 518
        -> font size < 8.5  => Result column
        -> font size >= 8.5 => Total column   (Total bold/bigger font hai,
           isi wajah se Result aur Total ke text kabhi kabhi visually overlap
           karte hain PDF me - font-size se hi reliably separate hote hain)
    M.S. No   : x0 >= 518

Agar kisi doosre report ka layout thoda alag hua (column x-positions
shift), to ye thresholds tune karne padenge - COLUMN_BOUNDS neeche
badal sakte ho.
"""

import re
import pdfplumber
import pandas as pd

COLUMN_BOUNDS = {
    "sno_max": 45,
    "roll_max": 155,
    "name_max": 259,
    "sdw_max": 371,
    "rt_max": 518,
}
TOTAL_FONT_SIZE_THRESHOLD = 8.5
HEADER_ROW_TOP_CUTOFF = 112  # rows above this 'top' value are page header, skip


def get_row_groups(chars, gap=6):
    chars = sorted(chars, key=lambda c: c['top'])
    rows = []
    cur = []
    cur_top = None
    for c in chars:
        if cur_top is None or c['top'] - cur_top <= gap:
            cur.append(c)
            cur_top = c['top'] if cur_top is None else max(cur_top, c['top'])
        else:
            rows.append(cur)
            cur = [c]
            cur_top = c['top']
    if cur:
        rows.append(cur)
    return rows


def chars_to_text(chars, gap_threshold=2.0):
    chars = sorted(chars, key=lambda c: c['x0'])
    text = ""
    prev_x1 = None
    for c in chars:
        if prev_x1 is not None and c['x0'] - prev_x1 > gap_threshold:
            if not text.endswith(" "):
                text += " "
        text += c['text']
        prev_x1 = c['x1']
    return text.strip()


def extract_page(page):
    text = page.extract_text() or ""
    study_centre = ""
    reg_centre = ""
    m = re.search(r'STUDY CENTRE\s*:\s*(.*?)\s+REG\.\s*CENTRE\s*:\s*(.*)', text)
    if m:
        study_centre = m.group(1).strip()
        reg_centre = m.group(2).strip().split('\n')[0]
    else:
        m2 = re.search(r'STUDY CENTRE\s*:\s*(.*)', text)
        if m2:
            study_centre = m2.group(1).strip()

    chars = page.chars
    data_chars = [c for c in chars if c['top'] > HEADER_ROW_TOP_CUTOFF]
    rows = get_row_groups(data_chars, gap=6)

    records = []
    b = COLUMN_BOUNDS
    for row in rows:
        row = sorted(row, key=lambda c: c['x0'])
        sno_chars = [c for c in row if c['x0'] < b["sno_max"]]
        roll_chars = [c for c in row if b["sno_max"] <= c['x0'] < b["roll_max"]]
        name_chars = [c for c in row if b["roll_max"] <= c['x0'] < b["name_max"]]
        sdw_chars = [c for c in row if b["name_max"] <= c['x0'] < b["sdw_max"]]
        rt_chars = [c for c in row if b["sdw_max"] <= c['x0'] < b["rt_max"]]
        ms_chars = [c for c in row if c['x0'] >= b["rt_max"]]

        result_chars = [c for c in rt_chars if c['size'] < TOTAL_FONT_SIZE_THRESHOLD]
        total_chars = [c for c in rt_chars if c['size'] >= TOTAL_FONT_SIZE_THRESHOLD]

        sno = chars_to_text(sno_chars)
        roll = chars_to_text(roll_chars)
        name = chars_to_text(name_chars)
        sdw = chars_to_text(sdw_chars)
        result = chars_to_text(result_chars)
        total = chars_to_text(total_chars)
        msno = chars_to_text(ms_chars)

        if not sno or not re.match(r'^\d+$', sno):
            continue
        if not roll:
            continue

        records.append({
            'S.No': sno,
            'Roll No': roll,
            'Name': name,
            'S/D/W Of': sdw,
            'Result': result,
            'Total': total,
            'M.S. No': msno,
            'Study Centre': study_centre,
            'Reg Centre': reg_centre,
        })
    return records


def find_expected_total(pdf):
    """Look for 'TOTAL STUDENTS : N' summary line, usually on the last page."""
    for page in reversed(pdf.pages[-3:]):
        text = page.extract_text() or ""
        m = re.search(r'TOTAL STUDENTS\s*:\s*(\d+)', text)
        if m:
            return int(m.group(1))
    return None


def extract_pdf_to_dataframe(pdf_path, progress_callback=None):
    """
    Extracts all student records from the PDF.
    progress_callback(current_page, total_pages) is called after each page, if provided.
    Returns: (dataframe, info_dict)
        info_dict = {
            "pages": total_pages,
            "records": count,
            "expected_total": N or None,
            "match": True/False/None
        }
    """
    all_records = []
    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        expected_total = find_expected_total(pdf)
        for i, page in enumerate(pdf.pages):
            recs = extract_page(page)
            all_records.extend(recs)
            if progress_callback:
                progress_callback(i + 1, total_pages)

    df = pd.DataFrame(all_records, columns=[
        'S.No', 'Roll No', 'Name', 'S/D/W Of', 'Result', 'Total',
        'M.S. No', 'Study Centre', 'Reg Centre'
    ])
    if not df.empty:
        df['S.No'] = range(1, len(df) + 1)

    info = {
        "pages": total_pages,
        "records": len(df),
        "expected_total": expected_total,
        "match": (expected_total == len(df)) if expected_total is not None else None,
    }
    return df, info


def save_to_excel(df, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = list(df.columns)
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    for row in df.itertuples(index=False):
        ws.append(list(row))

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="center", horizontal="left" if c not in (1, 6) else "center")

    widths = {'A': 8, 'B': 16, 'C': 26, 'D': 26, 'E': 32, 'F': 12, 'G': 16, 'H': 42, 'I': 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
