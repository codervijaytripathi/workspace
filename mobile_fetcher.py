"""
MP Bhoj Open University - Student data fetcher
UG + PG support.

Uses Selenium with the official MP Bhoj migration form.
No CAPTCHA/authentication/rate-limit bypass is attempted.
"""

import re
import time
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
)
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service

URL = "https://mpbou.mponline.gov.in/portal/Services/BHOJ/BrochureFee/Migration.aspx"
WAIT_TIMEOUT = 20
POLL_INTERVAL = 0.5


def normalize_course_type(value):
    value = str(value or "UG").strip().upper()
    return "PG" if value == "PG" else "UG"


def xpath_literal(s):
    if "'" not in s:
        return f"'{s}'"
    if '"' not in s:
        return f'"{s}"'
    parts = s.split("'")
    return "concat('" + "', \"'\", '".join(parts) + "')"


def find_select_near_label(driver, label_text):
    lit = xpath_literal(label_text)
    xpaths = [
        f"//td[contains(normalize-space(.),{lit})]/following-sibling::td[1]//select",
        f"//td[contains(normalize-space(.),{lit})]/following::select[1]",
        f"//label[contains(normalize-space(.),{lit})]/following::select[1]",
        f"//*[contains(normalize-space(.),{lit})]/following::select[1]",
    ]
    for xp in xpaths:
        elems = driver.find_elements(By.XPATH, xp)
        if elems:
            return elems[0]
    raise NoSuchElementException(f"Select dropdown near label '{label_text}' not found")


def find_input_near_label(driver, label_text):
    lit = xpath_literal(label_text)
    xpaths = [
        f"//td[contains(normalize-space(.),{lit})]/following-sibling::td[1]//input",
        f"//td[contains(normalize-space(.),{lit})]/following::input[1]",
        f"//label[contains(normalize-space(.),{lit})]/following::input[1]",
        f"//*[contains(normalize-space(.),{lit})]/following::input[1]",
    ]
    for xp in xpaths:
        elems = driver.find_elements(By.XPATH, xp)
        if elems:
            return elems[0]
    raise NoSuchElementException(f"Input near label '{label_text}' not found")


def try_find_input(driver, labels):
    for label in labels:
        try:
            return find_input_near_label(driver, label)
        except Exception:
            continue
    return None


def try_find_select(driver, labels):
    for label in labels:
        try:
            return find_select_near_label(driver, label)
        except Exception:
            continue
    return None


def get_input_value(driver, labels):
    el = try_find_input(driver, labels)
    if not el:
        return ""
    try:
        return (el.get_attribute("value") or "").strip()
    except Exception:
        return ""


def get_select_value(driver, labels):
    el = try_find_select(driver, labels)
    if not el:
        return ""
    try:
        return Select(el).first_selected_option.text.strip()
    except Exception:
        return ""


def select_by_visible_text_contains(select_el, text_fragment):
    sel = Select(select_el)
    fragment = text_fragment.lower().strip()
    for option in sel.options:
        if fragment in option.text.lower():
            sel.select_by_visible_text(option.text)
            return option.text
    raise NoSuchElementException(
        f"Option containing '{text_fragment}' not found in dropdown"
    )


def _wait_for_postback(driver, old_ref, wait):
    try:
        wait.until(EC.staleness_of(old_ref))
    except TimeoutException:
        pass
    try:
        wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
    except TimeoutException:
        pass
    time.sleep(0.7)


def setup_driver(headless=False):
    options = Options()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--window-size=1400,1000")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def setup_form(driver, course_type="UG"):
    """
    Open the official form and select Course Type + Apply For.
    Course Type is now dynamic: UG or PG.
    """
    course_type = normalize_course_type(course_type)
    driver.get(URL)
    wait = WebDriverWait(driver, WAIT_TIMEOUT)

    course_select = find_select_near_label(driver, "Course Type")
    old_ref = course_select
    select_by_visible_text_contains(course_select, course_type)
    _wait_for_postback(driver, old_ref, wait)

    apply_select = find_select_near_label(driver, "Apply For")
    old_ref = apply_select
    select_by_visible_text_contains(apply_select, "No Objection Certificate")
    _wait_for_postback(driver, old_ref, wait)

    return course_type


def _fetch_one_impl(driver, enrollment_no):
    wait = WebDriverWait(driver, WAIT_TIMEOUT)

    enrollment_input = find_input_near_label(driver, "Enrollment No")
    enrollment_input.clear()
    enrollment_input.send_keys(str(enrollment_no).strip())

    old_ref = enrollment_input
    enrollment_input.send_keys("\t")

    try:
        wait.until(EC.staleness_of(old_ref))
    except TimeoutException:
        pass

    try:
        wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
    except TimeoutException:
        pass

    time.sleep(0.7)

    # The portal may populate several fields after Enrollment No.
    values = {
        "Enrollment No": str(enrollment_no).strip(),
        "Candidate Name": get_input_value(driver, ["Candidate's Name", "Candidate Name", "Name"]),
        "Father's Name": get_input_value(driver, ["Father's Name", "Father Name"]),
        "Mobile No": get_input_value(driver, ["Mobile No", "Mobile Number", "Mobile"]),
        "Date of Birth": get_input_value(driver, ["Date Of Birth", "Date of Birth", "DOB"]),
        "Regional Centre": get_select_value(driver, ["Regional Centre", "Regional Center"]),
        "Course": get_select_value(driver, ["Course"]),
        "Study Centre": get_input_value(driver, ["Study Centre", "Study Center"]),
    }

    # Some fields can be ordinary text inputs rather than selects.
    if not values["Regional Centre"]:
        values["Regional Centre"] = get_input_value(
            driver, ["Regional Centre", "Regional Center"]
        )
    if not values["Course"]:
        values["Course"] = get_input_value(driver, ["Course"])
    if not values["Study Centre"]:
        values["Study Centre"] = values["Regional Centre"]

    if values["Mobile No"]:
        values["Status"] = "Found"
    elif any(
        values[k]
        for k in ("Candidate Name", "Father's Name", "Date of Birth", "Course", "Regional Centre")
    ):
        values["Status"] = "Found"
    else:
        values["Status"] = "Not Found / No Response"

    return values


def fetch_one(driver, enrollment_no, max_retries=3):
    last_exc = None
    for _ in range(max_retries):
        try:
            return _fetch_one_impl(driver, enrollment_no)
        except StaleElementReferenceException as exc:
            last_exc = exc
            time.sleep(1.0)
    if last_exc:
        raise last_exc
    return _fetch_one_impl(driver, enrollment_no)


def filter_result(result, selected_fields=None):
    """
    Keep the internal fetch complete, but return only fields selected by UI.
    Enrollment No and Status are always retained for identification/progress.
    """
    if not selected_fields:
        return result

    selected = set(selected_fields)
    output = {
        "Enrollment No": result.get("Enrollment No", ""),
        "Status": result.get("Status", ""),
    }

    mapping = {
        "name": "Candidate Name",
        "father_name": "Father's Name",
        "mobile": "Mobile No",
        "enrollment": "Enrollment No",
        "course": "Course",
        "dob": "Date of Birth",
        "study_centre": "Study Centre",
    }

    for key, source_key in mapping.items():
        if key in selected:
            output[source_key] = result.get(source_key, "")

    return output


def test_single(enrollment_no, course_type="UG", selected_fields=None,
                headless=False, screenshot_path=None):
    driver = setup_driver(headless=headless)
    try:
        setup_form(driver, course_type=course_type)
        result = fetch_one(driver, enrollment_no)
        return filter_result(result, selected_fields)
    except Exception:
        if screenshot_path:
            try:
                driver.save_screenshot(str(screenshot_path))
            except Exception:
                pass
        raise
    finally:
        driver.quit()


def read_enrollment_numbers(input_path):
    import pandas as pd
    input_path = Path(input_path)
    suffix = input_path.suffix.lower()

    if suffix in (".xlsx", ".xls", ".csv"):
        if suffix == ".csv":
            df = pd.read_csv(input_path, dtype=str)
        else:
            df = pd.read_excel(input_path, dtype=str)

        if df.empty:
            return []

        target_col = None
        for col in df.columns:
            if re.search(r"enroll|roll", str(col), re.IGNORECASE):
                target_col = col
                break

        if target_col is None:
            target_col = df.columns[0]

        raw = df[target_col].dropna().astype(str).str.strip().tolist()
        numbers = []
        seen = set()

        for value in raw:
            if not value or value.lower() == "nan":
                continue
            # Avoid Excel converting integer-looking enrollment numbers to x.0
            value = re.sub(r"\.0$", "", value)
            if value not in seen:
                seen.add(value)
                numbers.append(value)
        return numbers

    if suffix == ".pdf":
        import pdfplumber

        text_parts = []
        with pdfplumber.open(input_path) as pdf:
            for page in pdf.pages:
                text_parts.append(page.extract_text() or "")

        text = "\n".join(text_parts)
        matches = re.findall(r"\b[A-Z]?\d{10,12}\b", text, re.IGNORECASE)

        numbers = []
        seen = set()
        for match in matches:
            value = match.strip()
            key = value.upper()
            if key not in seen:
                seen.add(key)
                numbers.append(value)
        return numbers

    raise ValueError("Unsupported file type. Use .xlsx, .xls, .csv or .pdf")


def save_results_to_excel(results, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Results"

    preferred = [
        "Course Type",
        "Enrollment No",
        "Candidate Name",
        "Father's Name",
        "Mobile No",
        "Date of Birth",
        "Course",
        "Study Centre",
        "Regional Centre",
        "Status",
    ]

    present = set()
    for row in results:
        present.update(row.keys())

    headers = [h for h in preferred if h in present]
    for row in results:
        for key in row.keys():
            if key not in headers:
                headers.append(key)

    if not headers:
        headers = ["Enrollment No", "Status"]

    ws.append(headers)

    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in results:
        ws.append([row.get(header, "") for header in headers])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        max_len = max(len(str(c.value or "")) for c in col_cells[:100])
        ws.column_dimensions[letter].width = min(max(max_len + 2, 14), 35)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
